"""
School Management System — Gayatri Vidyapeeth, Daudnagar
Production-grade Flask application with comprehensive security hardening.

Security features:
  - bcrypt password hashing with SHA-256 migration support
  - CSRF protection on all forms (Flask-WTF)
  - Rate limiting on login routes (Flask-Limiter)
  - Login attempt tracking with account lockout
  - Input validation and sanitization (bleach)
  - NoSQL injection prevention
  - Security headers (X-Frame-Options, X-Content-Type-Options, etc.)
  - Session security (HttpOnly, SameSite, timeout, regeneration)
  - Structured logging with rotation
  - Custom error handlers
  - ObjectId validation on all DB operations
  - POST-only destructive operations
"""

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, flash, send_file, abort
)
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect
from datetime import date, datetime, timezone, timedelta
from werkzeug.utils import secure_filename
from functools import wraps
import re
import secrets
import pandas as pd
import io
import calendar
import uuid
import os
import logging
from logging.handlers import RotatingFileHandler

from bson.objectid import ObjectId
from pymongo import MongoClient
from dotenv import load_dotenv

# Load environment variables FIRST
load_dotenv()

# Import security modules
from security import (
    SecurityValidator, PasswordManager, LoginAttemptTracker,
    requires_role, safe_str
)
from config import config
from middleware import SecurityMiddleware

# ─── App Factory ─────────────────────────────────────────────────────────────

app = Flask(__name__)

# Load configuration
env = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config.get(env, config['default']))

# ─── CSRF Protection ────────────────────────────────────────────────────────
csrf = CSRFProtect(app)

# ─── Rate Limiting ──────────────────────────────────────────────────────────
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "60 per hour"],
    storage_uri=app.config.get('RATELIMIT_STORAGE_URL', 'memory://')
)

# ─── Security Middleware ────────────────────────────────────────────────────
SecurityMiddleware(app)

# ─── Logging ────────────────────────────────────────────────────────────────
log_dir = os.path.dirname(app.config.get('LOG_FILE', 'logs/school_app.log'))
if log_dir and not os.path.exists(log_dir):
    os.makedirs(log_dir, exist_ok=True)

file_handler = RotatingFileHandler(
    app.config.get('LOG_FILE', 'logs/school_app.log'),
    maxBytes=10 * 1024 * 1024,  # 10 MB
    backupCount=10
)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [%(pathname)s:%(lineno)d]'
))
file_handler.setLevel(getattr(logging, app.config.get('LOG_LEVEL', 'INFO')))
app.logger.addHandler(file_handler)
app.logger.setLevel(getattr(logging, app.config.get('LOG_LEVEL', 'INFO')))
app.logger.info('School Management System starting up')

# ─── MongoDB Connection ─────────────────────────────────────────────────────
mongo_uri = app.config['MONGO_URI']
try:
    client = MongoClient(mongo_uri, serverSelectionTimeoutMS=5000)
    client.server_info()  # Test connection
    db = client['gayatri_school']
    app.logger.info('MongoDB connection successful')
except Exception as e:
    app.logger.critical(f'MongoDB connection failed: {e}')
    raise

# Database Collections
teachers_col = db['teachers']
attendance_col = db['attendance']
admins_col = db['admins']
principals_col = db['principals']
increment_col = db['increments']
holidays_col = db['govt_holidays']
logs_col = db['activity_logs']
assets_col = db['assets']
students_col = db['students']
fee_history_col = db['fee_history']
leave_requests_col = db['leave_requests']
certificates_col = db['certificates']

# ─── Accountant Blueprint ───────────────────────────────────────────────────
from accountant_bp import accountant_bp, init_accountant
app.register_blueprint(accountant_bp)

# ─── Flask-Mail ─────────────────────────────────────────────────────────────
mail = Mail(app)

# ─── Login Attempt Tracker ──────────────────────────────────────────────────
login_tracker = LoginAttemptTracker(
    db,
    max_attempts=app.config.get('MAX_LOGIN_ATTEMPTS', 5),
    lockout_duration_minutes=int(
        app.config.get('LOGIN_LOCKOUT_DURATION', timedelta(minutes=15)).total_seconds() / 60
    )
)

# ─── Upload Config ──────────────────────────────────────────────────────────
UPLOAD_FOLDER = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    app.config.get('UPLOAD_FOLDER', 'static/uploads')
)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = app.config.get(
    'ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'gif', 'webp'}
)

# ─── Database Indexes ───────────────────────────────────────────────────────
try:
    teachers_col.create_index('teacher_id', unique=True)
    teachers_col.create_index('phone')
    attendance_col.create_index([('teacher_id', 1), ('date', -1)])
    students_col.create_index('admission_no')
    students_col.create_index([('class', 1), ('section', 1)])
    fee_history_col.create_index([('student_id', 1), ('date', -1)])
    fee_history_col.create_index('receipt_no', unique=True)
    logs_col.create_index([('teacher_id', 1), ('timestamp', -1)])
    holidays_col.create_index('date', unique=True)
    leave_requests_col.create_index([('teacher_id', 1), ('applied_on', -1)])
    app.logger.info('Database indexes created/verified')
except Exception as e:
    app.logger.warning(f'Index creation warning: {e}')


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def log_activity(teacher_id, teacher_name, action, details=''):
    """Log teacher activity to MongoDB (sanitized)."""
    try:
        ist_now = datetime.now(timezone(timedelta(hours=5, minutes=30)))
        logs_col.insert_one({
            'teacher_id': SecurityValidator.sanitize_string(str(teacher_id), 50),
            'teacher_name': SecurityValidator.sanitize_string(str(teacher_name), 100),
            'action': SecurityValidator.sanitize_string(str(action), 100),
            'details': SecurityValidator.sanitize_string(str(details), 500),
            'ip': request.remote_addr,
            'user_agent': request.headers.get('User-Agent', '')[:500],
            'timestamp': ist_now,
            'date': ist_now.strftime('%Y-%m-%d'),
            'time': ist_now.strftime('%I:%M:%S %p')
        })
    except Exception as e:
        app.logger.error(f'Activity logging error: {e}')


def log_security_event(event_type, username, details=''):
    """Log security-related events."""
    try:
        app.logger.warning(
            'SECURITY: %s | User: %s | IP: %s | UA: %s | %s',
            event_type,
            SecurityValidator.sanitize_string(str(username), 50),
            request.remote_addr,
            request.headers.get('User-Agent', '')[:200],
            SecurityValidator.sanitize_string(str(details), 500)
        )
    except Exception:
        pass


# ─── Auth Decorators ────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin'):
            log_security_event('UNAUTH_ACCESS', 'anonymous', f'Path: {request.path}')
            flash('कृपया लॉगिन करें!')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def principal_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('principal') and not session.get('admin'):
            log_security_event('UNAUTH_ACCESS', 'anonymous', f'Path: {request.path}')
            flash('कृपया लॉगिन करें!')
            return redirect(url_for('principal_login'))
        return f(*args, **kwargs)
    return decorated_function


def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('teacher_id'):
            log_security_event('UNAUTH_ACCESS', 'anonymous', f'Path: {request.path}')
            flash('कृपया लॉगिन करें!')
            return redirect(url_for('teacher_login'))
        return f(*args, **kwargs)
    return decorated_function


def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('student_id'):
            flash('कृपया लॉगिन करें!')
            return redirect(url_for('student_login'))
        return f(*args, **kwargs)
    return decorated_function


# ─── Salary Calculation Helpers ─────────────────────────────────────────────

def get_salary_calculation_days(year, month):
    """
    Salary calculation ke liye base days return karta hai.
    Hamesha 30 — chahe month 28, 29, 30 ya 31 din ka ho.
    """
    return 30  # ALWAYS 30 days


def get_month_summary(year, month):
    """Returns complete month summary: total days, sundays, govt holidays, working days."""
    days_in_month = calendar.monthrange(year, month)[1]
    month_str = f"{year}-{month:02d}"

    # Count Sundays
    sundays = 0
    sunday_days = set()
    for d in range(1, days_in_month + 1):
        if calendar.weekday(year, month, d) == 6:
            sundays += 1
            sunday_days.add(d)

    # Get govt holidays from DB
    govt_holidays = list(holidays_col.find(
        {'date': {'$regex': f'^{re.escape(month_str)}'}}
    ).sort('date', 1))
    holiday_days = set()
    for h in govt_holidays:
        d = int(h['date'].split('-')[2])
        if d not in sunday_days:
            holiday_days.add(d)

    actual_working_days = days_in_month - sundays - len(holiday_days)
    salary_calc_days = get_salary_calculation_days(year, month)

    return {
        'days_in_month': days_in_month,
        'sundays': sundays,
        'sunday_days': sunday_days,
        'holidays': len(holiday_days),
        'holiday_days': holiday_days,
        'holidays_list': govt_holidays,
        'working_days': actual_working_days,
        'salary_calc_days': salary_calc_days
    }


def get_working_days(year, month):
    """Keep backward compatibility."""
    return get_month_summary(year, month)['working_days']


def detect_continuous_leave_periods(tid, year, month, sunday_days=None):
    """
    Detect continuous leave periods (3+ consecutive absent days).
    Sundays are automatically included if they fall between absent days.
    """
    month_str = f"{year}-{month:02d}"
    days_in_month = calendar.monthrange(year, month)[1]

    if sunday_days is None:
        sunday_days = set()

    absent_records = list(attendance_col.find({
        'teacher_id': tid,
        'date': {'$regex': f'^{re.escape(month_str)}'},
        'status': {'$in': ['absent', 'A']}
    }).sort('date', 1))

    absent_days = sorted([int(rec['date'].split('-')[2]) for rec in absent_records])

    if not absent_days:
        return []

    expanded_absent = set(absent_days)
    for day in absent_days:
        for offset in [1, 2, 3]:
            next_day = day + offset
            if next_day in sunday_days and next_day <= days_in_month:
                if any(ad > next_day and ad <= next_day + 3 for ad in absent_days):
                    expanded_absent.add(next_day)

    absent_days = sorted(list(expanded_absent))

    continuous_periods = []
    current_start = absent_days[0]
    current_end = absent_days[0]

    for i in range(1, len(absent_days)):
        if absent_days[i] == current_end + 1:
            current_end = absent_days[i]
        else:
            if current_end - current_start + 1 >= 3:
                continuous_periods.append((current_start, current_end))
            current_start = absent_days[i]
            current_end = absent_days[i]

    if current_end - current_start + 1 >= 3:
        continuous_periods.append((current_start, current_end))

    return continuous_periods


def calculate_paid_days(tid, year, month, summary):
    """
    Attendance-based paid days calculation with Continuous Leave Rule.
    """
    month_str = f"{year}-{month:02d}"
    salary_calc_days = summary.get('salary_calc_days', 30)
    escaped_month = re.escape(month_str)

    present = attendance_col.count_documents({
        'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'},
        'status': {'$in': ['present', 'P']}
    })
    half = attendance_col.count_documents({
        'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'},
        'status': {'$in': ['half_day', 'H']}
    })
    medical = attendance_col.count_documents({
        'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'},
        'status': 'M'
    })
    absent = attendance_col.count_documents({
        'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'},
        'status': {'$in': ['absent', 'A']}
    })

    has_any_attendance = (present + half + medical) > 0

    if has_any_attendance:
        sunday_days = summary.get('sunday_days', set())
        holiday_days = summary.get('holiday_days', set())
        continuous_leave_periods = detect_continuous_leave_periods(
            tid, year, month, sunday_days
        )

        sundays_paid = 4
        holidays_paid = len(holiday_days)
        sundays_in_attendance = False

        # Check for salary adjustments in database
        salary_adj = db['salary_adjustments'].find_one({
            'teacher_id': tid, 'year': year, 'month': month
        })
        if salary_adj:
            sundays_paid = salary_adj.get('sundays_paid', sundays_paid)
            sundays_in_attendance = salary_adj.get('sundays_in_attendance', False)
        else:
            # Apply Continuous Leave Rule
            sundays_in_leave = 0
            holidays_in_leave = 0

            for leave_start, leave_end in continuous_leave_periods:
                for day in range(leave_start, leave_end + 1):
                    if day in sunday_days:
                        sundays_in_leave += 1
                    if day in holiday_days:
                        holidays_in_leave += 1

            sundays_paid = max(0, sundays_paid - sundays_in_leave)
            holidays_paid = max(0, holidays_paid - holidays_in_leave)

        if sundays_in_attendance:
            paid_days = present + medical + (half * 0.5) + holidays_paid
        else:
            paid_days = present + medical + (half * 0.5) + sundays_paid + holidays_paid

        paid_days = min(paid_days, salary_calc_days)
    else:
        sundays_paid = 0
        holidays_paid = 0
        paid_days = 0

    return {
        'present': present,
        'half': half,
        'medical': medical,
        'absent': absent,
        'sundays_paid': sundays_paid,
        'holidays_paid': holidays_paid,
        'paid_days': round(paid_days, 2),
        'leave_taken': absent,
    }


def compute_net_salary(basic_salary, att, salary_calc_days):
    """
    Net salary compute using PAID DAYS method.
    Formula: per_day = basic_salary / 30, net = per_day * paid_days
    """
    per_day = basic_salary / salary_calc_days if salary_calc_days > 0 else 0
    paid_days = att.get('paid_days', 0)
    net_salary = round(per_day * paid_days, 2)
    deduction = round(basic_salary - net_salary, 2)
    return net_salary, deduction, round(per_day, 2)


# ─── Account Initialization ────────────────────────────────────────────────

def init_admin():
    """Initialize default admin/principal accounts with bcrypt passwords."""
    pm = PasswordManager()

    # Admin
    admin_user = app.config['ADMIN_USERNAME']
    admin_pass = app.config['ADMIN_DEFAULT_PASSWORD']

    existing = admins_col.find_one({'username': admin_user})
    if not existing:
        admins_col.insert_one({
            'username': admin_user,
            'password': pm.hash_password(admin_pass),
            'name': 'Ravindra kumar',
            'created_at': datetime.now(timezone.utc),
            'must_change_password': True
        })
        app.logger.info(f'Admin account created: {admin_user}')
    elif pm.needs_rehash(existing.get('password', '')):
        # Only rehash if we can verify the old password
        if PasswordManager.verify_password(admin_pass, existing.get('password', '')):
            admins_col.update_one(
                {'username': admin_user},
                {'$set': {
                    'password': pm.hash_password(admin_pass),
                    'name': 'Ravindra kumar',
                    'migrated_at': datetime.now(timezone.utc)
                }}
            )
            app.logger.info(f'Admin password migrated to bcrypt: {admin_user}')

    # Principal
    prin_user = app.config['PRINCIPAL_USERNAME']
    prin_pass = app.config['PRINCIPAL_DEFAULT_PASSWORD']

    existing = principals_col.find_one({'username': prin_user})
    if not existing:
        principals_col.insert_one({
            'username': prin_user,
            'password': pm.hash_password(prin_pass),
            'name': 'Shivani singh',
            'created_at': datetime.now(timezone.utc),
            'must_change_password': True
        })
        app.logger.info(f'Principal account created: {prin_user}')
    elif pm.needs_rehash(existing.get('password', '')):
        if PasswordManager.verify_password(prin_pass, existing.get('password', '')):
            principals_col.update_one(
                {'username': prin_user},
                {'$set': {
                    'password': pm.hash_password(prin_pass),
                    'name': 'Shivani singh',
                    'migrated_at': datetime.now(timezone.utc)
                }}
            )
            app.logger.info(f'Principal password migrated to bcrypt: {prin_user}')


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Public
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Authentication
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def admin_login():
    if session.get('admin'):
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        username = safe_str(request.form.get('username', ''), 50).strip()
        password = safe_str(request.form.get('password', ''), 128)

        if not username or not password:
            flash('कृपया username और password दर्ज करें!')
            return render_template('admin_login.html')

        # Check lockout
        if login_tracker.is_locked(username):
            log_security_event('LOCKED_ACCOUNT', username, 'Login attempt on locked account')
            flash('बहुत अधिक असफल प्रयास! 15 मिनट बाद प्रयास करें।')
            return render_template('admin_login.html')

        admin = admins_col.find_one({'username': username})

        if admin and PasswordManager.verify_password(password, admin.get('password', '')):
            # Migrate password to bcrypt if needed
            if PasswordManager.needs_rehash(admin.get('password', '')):
                admins_col.update_one(
                    {'_id': admin['_id']},
                    {'$set': {'password': PasswordManager.hash_password(password)}}
                )

            login_tracker.record_attempt(username, success=True)
            login_tracker.reset_attempts(username)

            # Regenerate session
            session.clear()
            session['admin'] = True
            session['admin_name'] = admin.get('name', 'Admin')
            session.permanent = True

            app.logger.info(f'Admin login: {username} from {request.remote_addr}')
            return redirect(url_for('admin_dashboard'))

        # Failed login — generic message (prevents account enumeration)
        login_tracker.record_attempt(username, success=False)
        remaining = login_tracker.get_remaining_attempts(username)
        log_security_event('FAILED_LOGIN', username, f'Admin login failed. Remaining: {remaining}')
        flash('गलत username या password!')

    return render_template('admin_login.html')


@app.route('/principal/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def principal_login():
    if session.get('principal'):
        return redirect(url_for('principal_dashboard'))

    if request.method == 'POST':
        username = safe_str(request.form.get('username', ''), 50).strip()
        password = safe_str(request.form.get('password', ''), 128)

        if not username or not password:
            flash('कृपया username और password दर्ज करें!')
            return render_template('principal_login.html')

        if login_tracker.is_locked(username):
            flash('बहुत अधिक असफल प्रयास! 15 मिनट बाद प्रयास करें।')
            return render_template('principal_login.html')

        principal = principals_col.find_one({'username': username})

        if principal and PasswordManager.verify_password(password, principal.get('password', '')):
            if PasswordManager.needs_rehash(principal.get('password', '')):
                principals_col.update_one(
                    {'_id': principal['_id']},
                    {'$set': {'password': PasswordManager.hash_password(password)}}
                )

            login_tracker.record_attempt(username, success=True)
            login_tracker.reset_attempts(username)

            session.clear()
            session['principal'] = True
            session['principal_name'] = principal.get('name', 'Principal')
            session.permanent = True

            app.logger.info(f'Principal login: {username}')
            return redirect(url_for('principal_dashboard'))

        login_tracker.record_attempt(username, success=False)
        flash('गलत username या password!')

    return render_template('principal_login.html')


@app.route('/teacher/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def teacher_login():
    if session.get('teacher_id'):
        return redirect(url_for('teacher_dashboard'))

    if request.method == 'POST':
        teacher_id = safe_str(request.form.get('teacher_id', ''), 20).strip().upper()
        password = safe_str(request.form.get('password', ''), 128)

        if not teacher_id or not password:
            flash('कृपया ID और password दर्ज करें!')
            return render_template('teacher_login.html')

        # Validate teacher_id format
        valid, result = SecurityValidator.validate_teacher_id(teacher_id)
        if not valid:
            flash('गलत ID format!')
            return render_template('teacher_login.html')
        teacher_id = result

        if login_tracker.is_locked(teacher_id):
            flash('बहुत अधिक असफल प्रयास! 15 मिनट बाद प्रयास करें।')
            return render_template('teacher_login.html')

        teacher = teachers_col.find_one({'teacher_id': teacher_id})

        if teacher and PasswordManager.verify_password(password, teacher.get('password', '')):
            if PasswordManager.needs_rehash(teacher.get('password', '')):
                teachers_col.update_one(
                    {'_id': teacher['_id']},
                    {'$set': {'password': PasswordManager.hash_password(password)}}
                )

            login_tracker.record_attempt(teacher_id, success=True)
            login_tracker.reset_attempts(teacher_id)

            session.clear()
            session['teacher_id'] = teacher_id
            session['teacher_name'] = teacher['name']
            session.permanent = True

            log_activity(teacher_id, teacher['name'], 'LOGIN', 'Teacher logged in')

            if teacher.get('must_change_password'):
                flash('सुरक्षा के लिए कृपया अपना पासवर्ड बदलें।')
                return redirect(url_for('teacher_change_password'))
            return redirect(url_for('teacher_dashboard'))

        login_tracker.record_attempt(teacher_id, success=False)
        flash('गलत ID या password!')

    return render_template('teacher_login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Admin Dashboard
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    total_teachers = teachers_col.count_documents({'active': True})
    today_date = date.today()
    today_str = today_date.strftime('%Y-%m-%d')
    today_attendance = attendance_col.count_documents({
        'date': today_str, 'status': {'$in': ['present', 'P']}
    })
    absent_today = attendance_col.count_documents({
        'date': today_str, 'status': {'$in': ['absent', 'A']}
    })
    teachers = list(teachers_col.find({'active': True}))

    # Chart Data
    trend_labels = []
    trend_presents = []
    trend_absents = []

    for i in range(6, -1, -1):
        d = today_date - timedelta(days=i)
        d_str = d.strftime('%Y-%m-%d')
        trend_labels.append(d.strftime('%d %b'))
        p = attendance_col.count_documents({
            'date': d_str, 'status': {'$in': ['present', 'P']}
        })
        a = attendance_col.count_documents({
            'date': d_str, 'status': {'$in': ['absent', 'A']}
        })
        trend_presents.append(p)
        trend_absents.append(a)

    subject_counts = {}
    for t in teachers:
        subj = t.get('subject', 'Other') or 'Other'
        subject_counts[subj] = subject_counts.get(subj, 0) + 1

    pie_labels = list(subject_counts.keys())
    pie_data = list(subject_counts.values())

    return render_template('admin_dashboard.html',
                         total=total_teachers,
                         present_today=today_attendance,
                         absent_today=absent_today,
                         teachers=teachers,
                         today=today_str,
                         admin_name=session.get('admin_name'),
                         trend_labels=trend_labels,
                         trend_presents=trend_presents,
                         trend_absents=trend_absents,
                         pie_labels=pie_labels,
                         pie_data=pie_data,
                         total_students=students_col.count_documents({}),
                         total_balance=next(iter(students_col.aggregate([
                             {'$group': {'_id': None, 'bal': {'$sum': '$balance_fee'}}}
                         ])), {}).get('bal', 0))


@app.route('/principal/dashboard')
@principal_required
def principal_dashboard():
    total_teachers = teachers_col.count_documents({'active': True})
    today_str = date.today().strftime('%Y-%m-%d')
    today_attendance = attendance_col.count_documents({
        'date': today_str, 'status': {'$in': ['present', 'P']}
    })
    absent_today = attendance_col.count_documents({
        'date': today_str, 'status': {'$in': ['absent', 'A']}
    })
    return render_template('principal_dashboard.html',
                         total=total_teachers,
                         present_today=today_attendance,
                         absent_today=absent_today,
                         today=today_str,
                         principal_name=session.get('principal_name'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Teacher Management (Admin)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/teachers')
@admin_required
def manage_teachers():
    teachers = list(teachers_col.find({'active': True}))
    return render_template('manage_teachers.html', teachers=teachers)


@app.route('/admin/teacher/add', methods=['GET', 'POST'])
@admin_required
def add_teacher():
    if request.method == 'POST':
        name = SecurityValidator.sanitize_string(request.form.get('name', ''), 100)
        phone = safe_str(request.form.get('phone', ''), 15).strip()

        # Validate required fields
        if not name:
            flash('⚠️ नाम अनिवार्य है!')
            return redirect(url_for('add_teacher'))

        valid, phone_result = SecurityValidator.validate_phone(phone)
        if not valid:
            flash(f'⚠️ {phone_result}')
            return redirect(url_for('add_teacher'))
        phone = phone_result

        # Generate teacher ID
        if len(phone) >= 4:
            base_id = f"TCH{phone[-4:]}"
        else:
            base_id = f"TCH{phone.zfill(4)}"

        teacher_id = base_id
        if teachers_col.find_one({'teacher_id': teacher_id}):
            count = 1
            while teachers_col.find_one({'teacher_id': f"{base_id}-{count}"}):
                count += 1
            teacher_id = f"{base_id}-{count}"

        # Validate salary
        valid, salary_result = SecurityValidator.validate_amount(
            request.form.get('basic_salary', 0)
        )
        if not valid:
            flash(f'⚠️ Salary: {salary_result}')
            return redirect(url_for('add_teacher'))

        # Validate email if provided
        email = safe_str(request.form.get('email', ''), 255).strip()
        if email:
            valid, email_result = SecurityValidator.validate_email(email)
            if not valid:
                flash(f'⚠️ {email_result}')
                return redirect(url_for('add_teacher'))
            email = email_result

        default_password = app.config.get('DEFAULT_TEACHER_PASSWORD', 'GVP@2026')

        teacher = {
            'teacher_id': teacher_id,
            'name': name,
            'subject': SecurityValidator.sanitize_string(
                request.form.get('subject', ''), 50
            ),
            'phone': phone,
            'email': email,
            'basic_salary': salary_result,
            'password': PasswordManager.hash_password(default_password),
            'joining_date': safe_str(request.form.get('joining_date', ''), 10),
            'active': True,
            'created_at': datetime.now(timezone.utc),
            'must_change_password': True,
            'bank_name': SecurityValidator.sanitize_string(
                request.form.get('bank_name', ''), 100
            ),
            'bank_account': SecurityValidator.sanitize_string(
                request.form.get('bank_account', ''), 30
            ),
            'ifsc': safe_str(request.form.get('ifsc', ''), 11).upper(),
            'holder_name': SecurityValidator.sanitize_string(
                request.form.get('holder_name', ''), 100
            ),
            'pan_no': safe_str(request.form.get('pan_no', ''), 10).upper()
        }
        teachers_col.insert_one(teacher)
        app.logger.info(f'Teacher added: {teacher_id} by admin')
        flash(f'Teacher {name} सफलतापूर्वक जोड़े गए! ID: {teacher_id}')
        return redirect(url_for('manage_teachers'))

    return render_template('add_teacher.html')


@app.route('/admin/teacher/delete/<teacher_id>', methods=['POST'])
@admin_required
def delete_teacher(teacher_id):
    """Soft-delete teacher — POST only."""
    teacher_id = safe_str(teacher_id, 20).strip()
    valid, _ = SecurityValidator.validate_teacher_id(teacher_id)
    if not valid:
        flash('Invalid teacher ID!')
        return redirect(url_for('manage_teachers'))

    teachers_col.update_one({'teacher_id': teacher_id}, {'$set': {'active': False}})
    app.logger.info(f'Teacher deactivated: {teacher_id}')
    flash('Teacher हटा दिए गए!')
    return redirect(url_for('manage_teachers'))


@app.route('/admin/teacher/edit/<teacher_id>', methods=['GET', 'POST'])
@admin_required
def edit_teacher(teacher_id):
    teacher_id = safe_str(teacher_id, 20).strip()
    teacher = teachers_col.find_one({'teacher_id': teacher_id})
    if not teacher:
        flash('Teacher नहीं मिले!')
        return redirect(url_for('manage_teachers'))

    if request.method == 'POST':
        # Validate salary
        valid, salary_result = SecurityValidator.validate_amount(
            request.form.get('basic_salary', 0)
        )
        if not valid:
            flash(f'⚠️ {salary_result}')
            return redirect(url_for('edit_teacher', teacher_id=teacher_id))

        updates = {
            'name': SecurityValidator.sanitize_string(
                request.form.get('name', ''), 100
            ),
            'subject': SecurityValidator.sanitize_string(
                request.form.get('subject', ''), 50
            ),
            'phone': safe_str(request.form.get('phone', ''), 15).strip(),
            'email': safe_str(request.form.get('email', ''), 255).strip(),
            'basic_salary': salary_result,
            'joining_date': safe_str(request.form.get('joining_date', ''), 10),
            'bank_name': SecurityValidator.sanitize_string(
                request.form.get('bank_name', ''), 100
            ),
            'bank_account': SecurityValidator.sanitize_string(
                request.form.get('bank_account', ''), 30
            ),
            'ifsc': safe_str(request.form.get('ifsc', ''), 11).upper(),
            'holder_name': SecurityValidator.sanitize_string(
                request.form.get('holder_name', ''), 100
            ),
            'pan_no': safe_str(request.form.get('pan_no', ''), 10).upper()
        }
        teachers_col.update_one({'teacher_id': teacher_id}, {'$set': updates})
        flash(f'✅ {updates["name"]} की जानकारी सफलतापूर्वक अपडेट हो गई!')
        return redirect(url_for('manage_teachers'))

    return render_template('edit_teacher.html', teacher=teacher)


@app.route('/admin/teacher/reset_password/<teacher_id>', methods=['POST'])
@admin_required
def admin_reset_teacher_password(teacher_id):
    """Reset teacher password — POST only."""
    teacher_id = safe_str(teacher_id, 20).strip()
    teacher = teachers_col.find_one({'teacher_id': teacher_id})
    if not teacher:
        flash('Teacher नहीं मिले!')
        return redirect(url_for('manage_teachers'))

    default_password = app.config.get('DEFAULT_TEACHER_PASSWORD', 'GVP@2026')
    teachers_col.update_one(
        {'teacher_id': teacher_id},
        {'$set': {
            'password': PasswordManager.hash_password(default_password),
            'must_change_password': True
        }}
    )
    app.logger.info(f'Teacher password reset: {teacher_id}')
    flash(f'🔑 {teacher["name"]} का Password Reset हो गया! Default Password: {default_password}')
    return redirect(url_for('manage_teachers'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Attendance
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/attendance', methods=['GET', 'POST'])
@principal_required
def mark_attendance():
    selected_date = safe_str(request.args.get('date', date.today().strftime('%Y-%m-%d')), 10)

    # Validate date format
    valid, _ = SecurityValidator.validate_date(selected_date)
    if not valid:
        selected_date = date.today().strftime('%Y-%m-%d')

    teachers = list(teachers_col.find({'active': True}))

    existing = {}
    for rec in attendance_col.find({'date': selected_date}):
        existing[rec['teacher_id']] = rec['status']

    approved_leaves = leave_requests_col.find({
        'status': 'Approved',
        'start_date': {'$lte': selected_date},
        'end_date': {'$gte': selected_date}
    })
    teachers_on_leave = set([req['teacher_id'] for req in approved_leaves])

    if request.method == 'POST':
        att_date = safe_str(request.form.get('att_date', ''), 10)
        valid, _ = SecurityValidator.validate_date(att_date)
        if not valid:
            flash('⚠️ Invalid date!')
            return redirect(url_for('mark_attendance'))

        for teacher in teachers:
            tid = teacher['teacher_id']
            status = safe_str(request.form.get(f'status_{tid}', 'none'), 20)

            # Validate status value
            if status not in ('none', 'present', 'P', 'absent', 'A', 'half_day', 'H', 'M'):
                continue

            if status == 'none':
                attendance_col.delete_one({'teacher_id': tid, 'date': att_date})
            else:
                attendance_col.update_one(
                    {'teacher_id': tid, 'date': att_date},
                    {'$set': {
                        'teacher_id': tid,
                        'teacher_name': teacher['name'],
                        'date': att_date,
                        'status': status,
                        'marked_by': 'Admin' if session.get('admin') else 'Principal',
                        'marked_at': datetime.now(timezone.utc)
                    }},
                    upsert=True
                )
        flash(f'{att_date} की attendance सफलतापूर्वक save हो गई!')
        return redirect(url_for('mark_attendance', date=att_date))

    return render_template('mark_attendance.html',
                         teachers=teachers,
                         selected_date=selected_date,
                         existing=existing,
                         teachers_on_leave=teachers_on_leave)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Payroll
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/payroll')
@admin_required
def payroll():
    month = int(safe_str(request.args.get('month', date.today().month), 2) or date.today().month)
    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)

    # Bounds check
    month = max(1, min(12, month))
    year = max(2020, min(2100, year))

    teachers = list(teachers_col.find({'active': True}))
    summary = get_month_summary(year, month)
    working_days = summary['working_days']

    payroll_data = []
    total_payable = 0

    for teacher in teachers:
        tid = teacher['teacher_id']
        att = calculate_paid_days(tid, year, month, summary)
        salary_calc_days = summary.get('salary_calc_days', 30)
        net_salary, deduction, per_day_salary = compute_net_salary(
            teacher['basic_salary'], att, salary_calc_days
        )
        total_payable += net_salary

        payroll_data.append({
            'teacher_id': tid,
            'name': teacher['name'],
            'subject': teacher.get('subject', ''),
            'basic_salary': teacher['basic_salary'],
            'days_in_month': summary['days_in_month'],
            'sundays': att['sundays_paid'],
            'holidays': att['holidays_paid'],
            'working_days': working_days,
            'present_days': att['present'],
            'half_days': att['half'],
            'medical_leaves': att['medical'],
            'absent_days': att['absent'],
            'paid_days': att['paid_days'],
            'per_day': round(per_day_salary, 2),
            'deduction': deduction,
            'net_salary': net_salary,
            'calculation_days': salary_calc_days
        })

    return render_template('payroll.html',
                         payroll=payroll_data,
                         month=month, year=year,
                         month_name=calendar.month_name[month],
                         working_days=working_days,
                         total_payable=round(total_payable, 2),
                         summary=summary)


@app.route('/admin/payroll/chart')
@admin_required
def payroll_chart():
    return render_template('payroll_chart_may2026.html')


@app.route('/admin/attendance/report')
@principal_required
def attendance_report():
    month = int(safe_str(request.args.get('month', date.today().month), 2) or date.today().month)
    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)
    month = max(1, min(12, month))
    year = max(2020, min(2100, year))

    month_str = f"{year}-{month:02d}"
    escaped_month = re.escape(month_str)
    teachers = list(teachers_col.find({'active': True}))
    days_in_month = calendar.monthrange(year, month)[1]

    sundays = set()
    for d in range(1, days_in_month + 1):
        if calendar.weekday(year, month, d) == 6:
            sundays.add(d)

    report = []
    for teacher in teachers:
        tid = teacher['teacher_id']
        att_map = {}
        for rec in attendance_col.find({
            'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'}
        }):
            day = int(rec['date'].split('-')[2])
            att_map[day] = rec['status']
        report.append({
            'name': teacher['name'],
            'teacher_id': tid,
            'att_map': att_map
        })

    submission_logs = list(attendance_col.find(
        {'date': {'$regex': f'^{escaped_month}'}}
    ).sort('marked_at', -1).limit(30))

    return render_template('attendance_report.html',
                         report=report,
                         month=month, year=year,
                         month_name=calendar.month_name[month],
                         days=days_in_month,
                         sundays=sundays,
                         submission_logs=submission_logs)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Holidays
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/holidays', methods=['GET', 'POST'])
@admin_required
def manage_holidays():
    if request.method == 'POST':
        hdate = safe_str(request.form.get('date', ''), 10)
        hname = SecurityValidator.sanitize_string(request.form.get('name', ''), 100)

        valid, _ = SecurityValidator.validate_date(hdate)
        if not valid:
            flash('⚠️ Invalid date format!')
            return redirect(url_for('manage_holidays'))

        if not hname:
            flash('⚠️ Holiday name required!')
            return redirect(url_for('manage_holidays'))

        if not holidays_col.find_one({'date': hdate}):
            holidays_col.insert_one({
                'date': hdate,
                'name': hname,
                'added_by': session.get('admin_name'),
                'added_at': datetime.now(timezone.utc)
            })
            flash(f'✅ {hdate} — "{hname}" छुट्टी add हो गई!')
        else:
            flash('⚠️ यह date पहले से registered है!')
        return redirect(url_for('manage_holidays'))

    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)
    year = max(2020, min(2100, year))
    all_holidays = list(holidays_col.find(
        {'date': {'$regex': f'^{re.escape(str(year))}'}}
    ).sort('date', 1))
    return render_template('manage_holidays.html', holidays=all_holidays, year=year)


@app.route('/admin/holidays/delete/<holiday_id>', methods=['POST'])
@admin_required
def delete_holiday(holiday_id):
    """Delete holiday — POST only."""
    valid, _ = SecurityValidator.validate_object_id(holiday_id)
    if not valid:
        flash('Invalid ID!')
        return redirect(url_for('manage_holidays'))
    holidays_col.delete_one({'_id': ObjectId(holiday_id)})
    flash('छुट्टी हटा दी गई!')
    return redirect(url_for('manage_holidays'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Salary Increment
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/salary/increment', methods=['GET', 'POST'])
@admin_required
def salary_increment():
    teachers = list(teachers_col.find({'active': True}))

    if request.method == 'POST':
        teacher_id = safe_str(request.form.get('teacher_id', ''), 20).strip()
        increment_type = safe_str(request.form.get('increment_type', ''), 10)
        remarks = SecurityValidator.sanitize_string(
            request.form.get('remarks', ''), 500
        )

        if increment_type not in ('fixed', 'percent'):
            flash('⚠️ Invalid increment type!')
            return redirect(url_for('salary_increment'))

        valid, inc_val = SecurityValidator.validate_amount(
            request.form.get('increment_value', 0)
        )
        if not valid:
            flash(f'⚠️ {inc_val}')
            return redirect(url_for('salary_increment'))

        teacher = teachers_col.find_one({'teacher_id': teacher_id})
        if not teacher:
            flash('Teacher नहीं मिले!')
            return redirect(url_for('salary_increment'))

        old_salary = teacher['basic_salary']
        if increment_type == 'percent':
            new_salary = round(old_salary * (1 + inc_val / 100), 2)
        else:
            new_salary = round(old_salary + inc_val, 2)

        teachers_col.update_one(
            {'teacher_id': teacher_id},
            {'$set': {'basic_salary': new_salary}}
        )
        increment_col.insert_one({
            'teacher_id': teacher_id,
            'teacher_name': teacher['name'],
            'old_salary': old_salary,
            'new_salary': new_salary,
            'increment_type': increment_type,
            'increment_value': inc_val,
            'remarks': remarks,
            'date': datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            'done_by': session.get('admin_name'),
            'year': datetime.now().year
        })
        app.logger.info(
            f'Salary increment: {teacher_id} {old_salary} -> {new_salary} by {session.get("admin_name")}'
        )
        diff = new_salary - old_salary
        flash(f'✅ {teacher["name"]} की Salary ₹{old_salary:,.0f} → ₹{new_salary:,.0f} (+₹{diff:,.0f})')
        return redirect(url_for('salary_increment'))

    history = list(increment_col.find().sort('date', -1).limit(30))
    return render_template('salary_increment.html', teachers=teachers, history=history)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Assets
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/assets', methods=['GET', 'POST'])
@principal_required
def manage_assets():
    teachers = list(teachers_col.find({'active': True}))

    if request.method == 'POST':
        teacher_id = safe_str(request.form.get('teacher_id', ''), 20).strip()
        item_name = SecurityValidator.sanitize_string(
            request.form.get('item_name', ''), 200
        )
        remarks = SecurityValidator.sanitize_string(
            request.form.get('remarks', ''), 500
        )

        valid, quantity = SecurityValidator.validate_positive_int(
            request.form.get('quantity', 1), 'Quantity', 1000
        )
        if not valid:
            flash(f'⚠️ {quantity}')
            return redirect(url_for('manage_assets'))

        if not item_name:
            flash('⚠️ Item name required!')
            return redirect(url_for('manage_assets'))

        teacher = teachers_col.find_one({'teacher_id': teacher_id})
        if teacher:
            ist_now = datetime.now(timezone(timedelta(hours=5, minutes=30)))
            assets_col.insert_one({
                'teacher_id': teacher_id,
                'teacher_name': teacher['name'],
                'item_name': item_name,
                'quantity': quantity,
                'remarks': remarks,
                'assigned_by': session.get('admin_name') or session.get('principal_name'),
                'date': ist_now.strftime('%Y-%m-%d'),
                'timestamp': ist_now
            })
            flash(f'✅ {teacher["name"]} को {quantity}x {item_name} असाइन किया गया!')
        else:
            flash('⚠️ Teacher नहीं मिला!')
        return redirect(url_for('manage_assets'))

    all_assets = list(assets_col.find().sort('timestamp', -1))
    return render_template('manage_assets.html', teachers=teachers, assets=all_assets)


@app.route('/admin/assets/delete/<asset_id>', methods=['POST'])
@principal_required
def delete_asset(asset_id):
    """Delete asset — POST only."""
    valid, _ = SecurityValidator.validate_object_id(asset_id)
    if not valid:
        flash('Invalid ID!')
        return redirect(url_for('manage_assets'))
    assets_col.delete_one({'_id': ObjectId(asset_id)})
    flash('असाइनमेंट सफलतापूर्वक हटा दिया गया!')
    return redirect(url_for('manage_assets'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Teacher Portal
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/teacher/dashboard')
@teacher_required
def teacher_dashboard():
    tid = session['teacher_id']
    teacher = teachers_col.find_one({'teacher_id': tid})
    if not teacher:
        session.clear()
        return redirect(url_for('teacher_login'))

    month = date.today().month
    year = date.today().year

    # Previous Month
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev_summary = get_month_summary(prev_year, prev_month)
    prev_att = calculate_paid_days(tid, prev_year, prev_month, prev_summary)
    prev_salary_calc_days = prev_summary.get('salary_calc_days', 30)
    prev_paid_days = prev_att['paid_days']
    prev_estimated_salary, _, prev_per_day = compute_net_salary(
        teacher['basic_salary'], prev_att, prev_salary_calc_days
    )

    summary = get_month_summary(year, month)
    working_days = summary['working_days']
    att = calculate_paid_days(tid, year, month, summary)
    salary_calc_days = summary.get('salary_calc_days', 30)
    paid_days = att['paid_days']
    estimated_salary, _, per_day = compute_net_salary(
        teacher['basic_salary'], att, salary_calc_days
    )

    recent = list(attendance_col.find({'teacher_id': tid}).sort('date', -1).limit(10))
    assigned_assets = list(assets_col.find({'teacher_id': tid}).sort('timestamp', -1))

    log_activity(tid, teacher['name'], 'VISIT_DASHBOARD', 'Visited teacher dashboard')

    return render_template('teacher_dashboard.html',
                         teacher=teacher,
                         present=att['present'], half=att['half'],
                         absent=att['absent'],
                         total_days=summary['days_in_month'],
                         calculation_days=None,
                         paid_days=paid_days,
                         per_day=per_day,
                         estimated_salary=estimated_salary,
                         month_name=calendar.month_name[month],
                         year=year,
                         prev_month=prev_month,
                         prev_month_name=calendar.month_name[prev_month],
                         prev_year=prev_year,
                         prev_estimated_salary=prev_estimated_salary,
                         prev_paid_days=prev_paid_days,
                         prev_per_day=prev_per_day,
                         recent=recent,
                         assets=assigned_assets)


@app.route('/teacher/salary')
@teacher_required
def teacher_salary():
    tid = session['teacher_id']
    teacher = teachers_col.find_one({'teacher_id': tid})
    if not teacher:
        session.clear()
        return redirect(url_for('teacher_login'))

    month = int(safe_str(request.args.get('month', date.today().month), 2) or date.today().month)
    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)
    month = max(1, min(12, month))
    year = max(2020, min(2100, year))

    summary = get_month_summary(year, month)
    month_str = f"{year}-{month:02d}"
    escaped_month = re.escape(month_str)

    today = date.today()
    total_any = attendance_col.count_documents({
        'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'}
    })
    no_att_data = total_any == 0
    is_current_month = (year == today.year and month == today.month)
    if no_att_data and is_current_month and not request.args.get('force'):
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        return redirect(url_for('teacher_salary', month=prev_month, year=prev_year))

    att = calculate_paid_days(tid, year, month, summary)
    salary_calc_days = summary.get('salary_calc_days', 30)
    net_salary, deduction, per_day = compute_net_salary(
        teacher['basic_salary'], att, salary_calc_days
    )

    all_teachers = list(teachers_col.find({'active': True}, {'teacher_id': 1}).sort('_id', 1))
    bill_index = next(
        (i + 1 for i, t in enumerate(all_teachers) if t['teacher_id'] == tid), 1
    )
    unique_bill_no = f"GVP-{year}-{month:02d}-{bill_index:03d}"
    slip_date = today.strftime('%d/%m/%Y')

    log_activity(tid, teacher['name'], 'VISIT_SALARY',
                f'Viewed salary slip for {calendar.month_name[month]} {year}')

    return render_template('salary_slip.html',
                         teacher=teacher,
                         month=month, year=year,
                         month_name=calendar.month_name[month],
                         summary=summary,
                         total_working_days=att['paid_days'],
                         calculation_days=salary_calc_days,
                         present=att['present'], half=att['half'],
                         medical=att['medical'], absent=att['absent'],
                         sundays_paid=att['sundays_paid'],
                         holidays_paid=att['holidays_paid'],
                         paid_days=att['paid_days'],
                         leave_taken=att['leave_taken'],
                         per_day=round(per_day, 2),
                         deduction=deduction,
                         net_salary=net_salary,
                         bill_no=unique_bill_no,
                         slip_date=slip_date,
                         no_att_data=no_att_data,
                         is_admin=False)


@app.route('/admin/salary/slip/<teacher_id>')
@admin_required
def admin_salary_slip(teacher_id):
    teacher_id = safe_str(teacher_id, 20).strip()
    teacher = teachers_col.find_one({'teacher_id': teacher_id})
    if not teacher:
        flash('Teacher नहीं मिले!')
        return redirect(url_for('payroll'))

    month = int(safe_str(request.args.get('month', date.today().month), 2) or date.today().month)
    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)
    month = max(1, min(12, month))
    year = max(2020, min(2100, year))

    month_str = f"{year}-{month:02d}"
    escaped_month = re.escape(month_str)
    summary = get_month_summary(year, month)

    today = date.today()
    total_any = attendance_col.count_documents({
        'teacher_id': teacher_id, 'date': {'$regex': f'^{escaped_month}'}
    })
    no_att_data = total_any == 0
    is_current_month = (year == today.year and month == today.month)
    if no_att_data and is_current_month and not request.args.get('force'):
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        return redirect(url_for('admin_salary_slip', teacher_id=teacher_id,
                                month=prev_month, year=prev_year))

    att = calculate_paid_days(teacher_id, year, month, summary)
    salary_calc_days = summary.get('salary_calc_days', 30)
    net_salary, deduction, per_day = compute_net_salary(
        teacher['basic_salary'], att, salary_calc_days
    )

    all_teachers = list(teachers_col.find({'active': True}, {'teacher_id': 1}).sort('_id', 1))
    bill_index = next(
        (i + 1 for i, t in enumerate(all_teachers) if t['teacher_id'] == teacher_id), 1
    )
    unique_bill_no = f"GVP-{year}-{month:02d}-{bill_index:03d}"
    slip_date = today.strftime('%d/%m/%Y')

    return render_template('salary_slip.html',
                         teacher=teacher,
                         month=month, year=year,
                         month_name=calendar.month_name[month],
                         summary=summary,
                         total_working_days=att['paid_days'],
                         calculation_days=salary_calc_days,
                         present=att['present'], half=att['half'],
                         medical=att['medical'], absent=att['absent'],
                         sundays_paid=att['sundays_paid'],
                         holidays_paid=att['holidays_paid'],
                         paid_days=att['paid_days'],
                         leave_taken=att['leave_taken'],
                         per_day=round(per_day, 2),
                         deduction=deduction,
                         net_salary=net_salary,
                         bill_no=unique_bill_no,
                         slip_date=slip_date,
                         no_att_data=no_att_data,
                         is_admin=True)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Teacher Leave
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/teacher/leave', methods=['GET', 'POST'])
@teacher_required
def teacher_leave():
    tid = session['teacher_id']
    if request.method == 'POST':
        start_date = safe_str(request.form.get('start_date', ''), 10)
        end_date = safe_str(request.form.get('end_date', ''), 10)
        reason = SecurityValidator.sanitize_string(
            request.form.get('reason', ''), 500
        )

        valid_s, _ = SecurityValidator.validate_date(start_date)
        valid_e, _ = SecurityValidator.validate_date(end_date)
        if not valid_s or not valid_e:
            flash('⚠️ Invalid dates!')
            return redirect(url_for('teacher_leave'))

        if not reason:
            flash('⚠️ कारण अनिवार्य है!')
            return redirect(url_for('teacher_leave'))

        teacher = teachers_col.find_one({'teacher_id': tid})

        leave_requests_col.insert_one({
            'teacher_id': tid,
            'teacher_name': teacher['name'],
            'start_date': start_date,
            'end_date': end_date,
            'reason': reason,
            'status': 'Pending',
            'applied_on': datetime.now(timezone.utc)
        })
        flash('छुट्टी का आवेदन सफलतापूर्वक भेज दिया गया है!')
        return redirect(url_for('teacher_leave'))

    leaves = list(leave_requests_col.find({'teacher_id': tid}).sort('applied_on', -1))
    today_str = date.today().strftime('%Y-%m-%d')
    return render_template('teacher_leave.html', leaves=leaves, today_str=today_str)


@app.route('/teacher/attendance/report')
@teacher_required
def teacher_attendance_report():
    tid = session['teacher_id']
    teacher = teachers_col.find_one({'teacher_id': tid})
    if not teacher:
        session.clear()
        return redirect(url_for('teacher_login'))

    month = int(safe_str(request.args.get('month', date.today().month), 2) or date.today().month)
    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)
    month = max(1, min(12, month))
    year = max(2020, min(2100, year))

    month_str = f"{year}-{month:02d}"
    escaped_month = re.escape(month_str)
    days_in_month = calendar.monthrange(year, month)[1]

    att_map = {}
    for rec in attendance_col.find({
        'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'}
    }):
        day = int(rec['date'].split('-')[2])
        att_map[day] = rec['status']

    sundays = set()
    for d in range(1, days_in_month + 1):
        if calendar.weekday(year, month, d) == 6:
            sundays.add(d)

    p = sum(1 for v in att_map.values() if v in ['present', 'P'])
    h = sum(1 for v in att_map.values() if v in ['half_day', 'H'])
    m = sum(1 for v in att_map.values() if v == 'M')
    a = sum(1 for v in att_map.values() if v in ['absent', 'A'])

    log_activity(tid, teacher['name'], 'VISIT_ATTENDANCE',
                f'Viewed attendance for {calendar.month_name[month]} {year}')

    return render_template('teacher_attendance_report.html',
                         teacher=teacher,
                         att_map=att_map,
                         month=month, year=year,
                         month_name=calendar.month_name[month],
                         days=days_in_month,
                         sundays=sundays,
                         p_count=p, h_count=h, m_count=m, a_count=a)


@app.route('/teacher/profile', methods=['GET', 'POST'])
@teacher_required
def teacher_profile():
    tid = session['teacher_id']
    teacher = teachers_col.find_one({'teacher_id': tid})
    if not teacher:
        session.clear()
        return redirect(url_for('teacher_login'))

    if request.method == 'POST':
        if 'photo' not in request.files:
            flash('कोई फ़ाइल नहीं चुनी!')
            return redirect(url_for('teacher_profile'))

        file = request.files['photo']
        if file.filename == '':
            flash('कोई फ़ाइल नहीं चुनी!')
            return redirect(url_for('teacher_profile'))

        # Use secure file validation with random filename
        valid, result = SecurityValidator.validate_file_upload(file, ALLOWED_EXTENSIONS)
        if valid:
            file.save(os.path.join(UPLOAD_FOLDER, result))
            teachers_col.update_one(
                {'teacher_id': tid}, {'$set': {'photo': result}}
            )
            log_activity(tid, teacher['name'], 'PHOTO_UPLOAD', 'Updated profile photo')
            flash('✅ Profile photo सफलतापूर्वक update हो गई!')
        else:
            flash(f'❌ {result}')
        return redirect(url_for('teacher_profile'))

    log_activity(tid, teacher['name'], 'VISIT_PROFILE', 'Visited profile page')
    return render_template('teacher_profile.html', teacher=teacher)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Password Management
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/teacher/forgot_password', methods=['GET', 'POST'])
@limiter.limit("3 per minute")
def teacher_forgot_password():
    if request.method == 'POST':
        teacher_id = safe_str(request.form.get('teacher_id', ''), 20).strip().upper()
        phone = safe_str(request.form.get('phone', ''), 15).strip()

        # Validate inputs
        valid, result = SecurityValidator.validate_teacher_id(teacher_id)
        if not valid:
            flash('गलत ID format!')
            return redirect(url_for('teacher_forgot_password'))
        teacher_id = result

        teacher = teachers_col.find_one({'teacher_id': teacher_id, 'phone': phone})

        if teacher:
            if not teacher.get('email'):
                flash('आपका Email रजिस्टर्ड नहीं है! कृपया एडमिन से ईमेल अपडेट करवाएं।')
                return redirect(url_for('teacher_forgot_password'))

            # Generate cryptographically secure OTP
            otp = PasswordManager.generate_otp(6)
            session['otp'] = otp
            session['otp_created'] = datetime.now(timezone.utc).isoformat()
            session['otp_attempts'] = 0
            session['reset_teacher_id'] = teacher_id

            try:
                msg = Message(
                    "Password Reset OTP - Gayatri Vidyapith",
                    recipients=[teacher['email']]
                )
                msg.body = (
                    f"Hello {teacher['name']},\n\n"
                    f"Your OTP for password reset is: {otp}\n\n"
                    f"This OTP expires in {app.config.get('OTP_EXPIRY_MINUTES', 10)} minutes.\n"
                    f"Do not share this with anyone."
                )
                mail.send(msg)
                flash(f'एक OTP आपकी ईमेल पर भेज दिया गया है।')
                return redirect(url_for('teacher_verify_otp'))
            except Exception as e:
                app.logger.error(f'Mail error: {e}')
                flash('ईमेल भेजने में गड़बड़ हुई! कृपया बाद में प्रयास करें।')
        else:
            # Generic message — prevents account/phone enumeration
            flash('गलत ID या Phone Number!')

    return render_template('teacher_forgot_password.html')


@app.route('/teacher/verify_otp', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def teacher_verify_otp():
    if not session.get('reset_teacher_id') or not session.get('otp'):
        return redirect(url_for('teacher_forgot_password'))

    if request.method == 'POST':
        entered_otp = safe_str(request.form.get('otp', ''), 6)

        # Check OTP expiration
        otp_created = session.get('otp_created', '')
        if otp_created:
            created_time = datetime.fromisoformat(otp_created)
            expiry = timedelta(minutes=app.config.get('OTP_EXPIRY_MINUTES', 10))
            if datetime.now(timezone.utc) - created_time > expiry:
                session.pop('otp', None)
                session.pop('reset_teacher_id', None)
                flash('OTP expired! कृपया नया OTP प्राप्त करें।')
                return redirect(url_for('teacher_forgot_password'))

        # Check attempt limit
        otp_attempts = session.get('otp_attempts', 0)
        max_attempts = app.config.get('OTP_MAX_ATTEMPTS', 3)
        if otp_attempts >= max_attempts:
            session.pop('otp', None)
            session.pop('reset_teacher_id', None)
            flash('बहुत अधिक गलत प्रयास! कृपया नया OTP प्राप्त करें।')
            return redirect(url_for('teacher_forgot_password'))

        # Timing-safe comparison
        import hmac
        if hmac.compare_digest(entered_otp, session.get('otp', '')):
            session['otp_verified'] = True
            flash('OTP सत्यापित! अब अपना नया पासवर्ड सेट करें।')
            return redirect(url_for('teacher_reset_password'))

        session['otp_attempts'] = otp_attempts + 1
        flash('गलत OTP! कृपया फिर से चेक करें।')

    return render_template('teacher_verify_otp.html')


@app.route('/teacher/reset_password', methods=['GET', 'POST'])
def teacher_reset_password():
    if not session.get('reset_teacher_id') or not session.get('otp_verified'):
        return redirect(url_for('teacher_forgot_password'))

    if request.method == 'POST':
        new_password = safe_str(request.form.get('new_password', ''), 128)
        confirm_password = safe_str(request.form.get('confirm_password', ''), 128)

        # Validate password strength
        valid, msg = SecurityValidator.validate_password(new_password)
        if not valid:
            flash(f'⚠️ {msg}')
            return render_template('teacher_reset_password.html')

        if new_password != confirm_password:
            flash('पासवर्ड मेल नहीं खाते!')
            return render_template('teacher_reset_password.html')

        teachers_col.update_one(
            {'teacher_id': session['reset_teacher_id']},
            {'$set': {
                'password': PasswordManager.hash_password(new_password),
                'must_change_password': False
            }}
        )

        # Clear OTP session data
        session.pop('reset_teacher_id', None)
        session.pop('otp', None)
        session.pop('otp_verified', None)
        session.pop('otp_created', None)

        app.logger.info(f'Password reset for teacher via OTP')
        flash('पासवर्ड सफलतापूर्वक बदल दिया गया है! अब आप लॉगिन कर सकते हैं।')
        return redirect(url_for('teacher_login'))

    return render_template('teacher_reset_password.html')


@app.route('/teacher/change_password', methods=['GET', 'POST'])
@teacher_required
def teacher_change_password():
    if request.method == 'POST':
        old_password = safe_str(request.form.get('old_password', ''), 128)
        new_password = safe_str(request.form.get('new_password', ''), 128)
        confirm_password = safe_str(request.form.get('confirm_password', ''), 128)

        teacher = teachers_col.find_one({'teacher_id': session['teacher_id']})

        if not teacher or not PasswordManager.verify_password(
            old_password, teacher.get('password', '')
        ):
            flash('पुराना पासवर्ड गलत है!')
            return render_template('teacher_change_password.html')

        if new_password != confirm_password:
            flash('नया पासवर्ड मेल नहीं खाता!')
            return render_template('teacher_change_password.html')

        # Validate new password strength
        valid, msg = SecurityValidator.validate_password(new_password)
        if not valid:
            flash(f'⚠️ {msg}')
            return render_template('teacher_change_password.html')

        teachers_col.update_one(
            {'teacher_id': session['teacher_id']},
            {'$set': {
                'password': PasswordManager.hash_password(new_password),
                'must_change_password': False
            }}
        )
        flash('पासवर्ड सफलतापूर्वक बदल दिया गया है!')
        return redirect(url_for('teacher_dashboard'))

    return render_template('teacher_change_password.html')


@app.route('/teacher/holidays')
@teacher_required
def teacher_holidays():
    holidays = list(holidays_col.find().sort('date', 1))
    return render_template('teacher_holidays.html', holidays=holidays)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Teacher Logs (Admin)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/teacher/logs')
@admin_required
def teacher_logs():
    filter_id = safe_str(request.args.get('teacher_id', ''), 20)
    filter_action = safe_str(request.args.get('action', ''), 50)
    filter_date = safe_str(request.args.get('date', ''), 10)

    query = {}
    if filter_id:
        query['teacher_id'] = filter_id
    if filter_action:
        query['action'] = filter_action
    if filter_date:
        query['date'] = filter_date

    logs = list(logs_col.find(query).sort('timestamp', -1).limit(200))
    all_teachers = list(teachers_col.find({'active': True}, {'teacher_id': 1, 'name': 1}))

    ist_now = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    today_ist = ist_now.strftime('%Y-%m-%d')
    today_logins = logs_col.count_documents({'action': 'LOGIN', 'date': today_ist})
    total_logins = logs_col.count_documents({'action': 'LOGIN'})
    total_visits = logs_col.count_documents({})

    return render_template('teacher_logs.html',
                         logs=logs,
                         all_teachers=all_teachers,
                         filter_id=filter_id,
                         filter_action=filter_action,
                         filter_date=filter_date,
                         today_logins=today_logins,
                         total_logins=total_logins,
                         total_visits=total_visits,
                         today=today_ist)


@app.route('/admin/teacher/logs/clear', methods=['POST'])
@admin_required
def clear_logs():
    ist_now = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    before_date = (ist_now - timedelta(days=2)).strftime('%Y-%m-%d')
    deleted_info = logs_col.delete_many({'date': {'$lt': before_date}})
    if deleted_info.deleted_count > 0:
        flash(f'✅ {before_date} से पहले के {deleted_info.deleted_count} logs delete हो गए!')
    else:
        flash(f'ℹ️ {before_date} से पहले के कोई logs मौजूद नहीं हैं।')
    return redirect(url_for('teacher_logs'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Attendance Export
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/attendance/export')
@principal_required
def export_attendance():
    month = int(safe_str(request.args.get('month', date.today().month), 2) or date.today().month)
    year = int(safe_str(request.args.get('year', date.today().year), 4) or date.today().year)
    month = max(1, min(12, month))
    year = max(2020, min(2100, year))

    month_str = f"{year}-{month:02d}"
    escaped_month = re.escape(month_str)
    teachers = list(teachers_col.find({'active': True}))
    days_in_month = calendar.monthrange(year, month)[1]

    data = []
    for teacher in teachers:
        tid = teacher['teacher_id']
        att_map = {}
        for rec in attendance_col.find({
            'teacher_id': tid, 'date': {'$regex': f'^{escaped_month}'}
        }):
            day = int(rec['date'].split('-')[2])
            att_map[day] = rec['status']

        row = {"Teacher Name": teacher['name'], "ID": tid}
        counts = {'P': 0, 'H': 0, 'M': 0, 'A': 0}
        for d in range(1, days_in_month + 1):
            s = att_map.get(d, "-")
            if s in ['present', 'P']:
                status = 'P'
                counts['P'] += 1
            elif s in ['half_day', 'H']:
                status = 'H'
                counts['H'] += 1
            elif s in ['absent', 'A']:
                status = 'A'
                counts['A'] += 1
            elif s == 'M':
                status = 'M'
                counts['M'] += 1
            else:
                status = "-"
            row[str(d)] = status

        row.update({
            "P (Present)": counts['P'],
            "H (Half Day)": counts['H'],
            "M (Medical)": counts['M'],
            "A (Absent)": counts['A']
        })
        data.append(row)

    df = pd.DataFrame(data)

    output = io.BytesIO()
    month_name = calendar.month_name[month]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Attendance_{month_str}', startrow=4)

        workbook = writer.book
        worksheet = writer.sheets[f'Attendance_{month_str}']

        from openpyxl.styles import Font, Alignment, PatternFill

        header_font = Font(bold=True, size=16, color="FFFFFF")
        sub_header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

        last_col = chr(ord("A") + min(days_in_month + 5, 25))
        worksheet.merge_cells(f'A1:{last_col}1')
        worksheet['A1'] = "गायत्री विद्यापीठ, दाउदनगर"
        worksheet['A1'].font = header_font
        worksheet['A1'].alignment = center_align
        worksheet['A1'].fill = header_fill

        worksheet.merge_cells(f'A2:{last_col}2')
        worksheet['A2'] = f"Attendance Report — {month_name} {year}"
        worksheet['A2'].font = sub_header_font
        worksheet['A2'].alignment = center_align

        worksheet.merge_cells(f'A3:{last_col}3')
        worksheet['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet['A3'].alignment = center_align

        for col in worksheet.columns:
            max_length = 0
            column = col[4].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column].width = max_length + 2

    output.seek(0)
    filename = f"Attendance_Report_{month_name}_{year}.xlsx"

    return send_file(output,
                     download_name=filename,
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Salary Slip Generator (Admin)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/salary/slip-generator', methods=['GET', 'POST'])
@admin_required
def salary_slip_generator():
    teachers = list(teachers_col.find({'active': True}))

    if request.method == 'POST':
        teacher_id = safe_str(request.form.get('teacher_id', ''), 20).strip()
        month = int(safe_str(request.form.get('month', 1), 2) or 1)
        year = int(safe_str(request.form.get('year', 2026), 4) or 2026)
        month = max(1, min(12, month))

        valid, present_days = SecurityValidator.validate_positive_int(
            request.form.get('present_days', 0), 'Present Days', 31
        )
        if not valid:
            flash(f'⚠️ {present_days}')
            return redirect(url_for('salary_slip_generator'))

        valid, absent_days = SecurityValidator.validate_positive_int(
            request.form.get('absent_days', 0), 'Absent Days', 31
        )
        valid2, paid_leave = SecurityValidator.validate_positive_int(
            request.form.get('paid_leave', 0), 'Paid Leave', 31
        )
        valid3, sunday_count = SecurityValidator.validate_positive_int(
            request.form.get('sunday_count', 4), 'Sunday Count', 10
        )

        teacher = teachers_col.find_one({'teacher_id': teacher_id})
        if not teacher:
            flash('Teacher नहीं मिले!')
            return redirect(url_for('salary_slip_generator'))

        basic_salary = teacher['basic_salary']
        salary_calc_days = 30
        paid_days = min(present_days + paid_leave + sunday_count, salary_calc_days)

        att = {
            'present': present_days,
            'half': 0,
            'medical': paid_leave,
            'absent': absent_days,
            'sundays_paid': sunday_count,
            'holidays_paid': 0,
            'paid_days': paid_days,
            'leave_taken': absent_days,
        }

        net_salary, deduction, per_day = compute_net_salary(
            basic_salary, att, salary_calc_days
        )

        all_teachers = list(teachers_col.find({'active': True}, {'teacher_id': 1}).sort('_id', 1))
        bill_index = next(
            (i + 1 for i, t in enumerate(all_teachers) if t['teacher_id'] == teacher_id), 1
        )
        unique_bill_no = f"GVP-SG-{year}-{month:02d}-{bill_index:03d}"
        slip_date = date.today().strftime('%d/%m/%Y')

        return render_template('salary_slip_generated.html',
                             teacher=teacher,
                             month=month, year=year,
                             month_name=calendar.month_name[month],
                             present_days=present_days,
                             absent_days=absent_days,
                             paid_leave=paid_leave,
                             sunday_count=sunday_count,
                             paid_days=paid_days,
                             basic_salary=basic_salary,
                             per_day=round(per_day, 2),
                             allowances=0,
                             deduction=deduction,
                             net_salary=net_salary,
                             bill_no=unique_bill_no,
                             slip_date=slip_date)

    today = date.today()
    return render_template('salary_slip_generator.html',
                         teachers=teachers,
                         current_month=today.month,
                         current_year=today.year)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Leave Requests (Admin)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/admin/leave/requests', methods=['GET', 'POST'])
@admin_required
def admin_leave_requests():
    if request.method == 'POST':
        req_id = safe_str(request.form.get('request_id', ''), 30)
        action = safe_str(request.form.get('action', ''), 10)

        valid, _ = SecurityValidator.validate_object_id(req_id)
        if not valid:
            flash('Invalid request ID!')
            return redirect(url_for('admin_leave_requests'))

        if action not in ('approve', 'reject'):
            flash('Invalid action!')
            return redirect(url_for('admin_leave_requests'))

        status = 'Approved' if action == 'approve' else 'Rejected'
        leave_requests_col.update_one(
            {'_id': ObjectId(req_id)}, {'$set': {'status': status}}
        )
        flash(f'Leave request {status}!')
        return redirect(url_for('admin_leave_requests'))

    requests_list = list(leave_requests_col.find().sort(
        [('status', -1), ('applied_on', -1)]
    ))
    return render_template('admin_leave_requests.html', requests=requests_list)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Student Fee Management (Admin)
# ═══════════════════════════════════════════════════════════════════════════

def get_ist_now_admin():
    return datetime.now(timezone(timedelta(hours=5, minutes=30)))


@app.route('/admin/students')
@admin_required
def manage_students():
    filter_class = safe_str(request.args.get('class', '')).strip()
    filter_section = safe_str(request.args.get('section', '')).strip()
    filter_search = safe_str(request.args.get('search', '')).strip()
    filter_fee_status = safe_str(request.args.get('fee_status', '')).strip()

    query = {}
    if filter_class:
        query['class'] = filter_class
    if filter_section:
        query['section'] = filter_section
    if filter_search:
        escaped = SecurityValidator.sanitize_search(filter_search)
        query['$or'] = [
            {'name': {'$regex': escaped, '$options': 'i'}},
            {'roll_no': {'$regex': escaped, '$options': 'i'}},
            {'admission_no': {'$regex': escaped, '$options': 'i'}},
        ]

    students = list(students_col.find(query).sort(
        [('class', 1), ('section', 1), ('roll_no', 1)]
    ))

    if filter_fee_status == 'fully_paid':
        students = [s for s in students if s.get('balance_fee', 0) <= 0]
    elif filter_fee_status == 'pending':
        students = [s for s in students
                    if s.get('paid_fee', 0) == 0 and s.get('total_fee', 0) > 0]
    elif filter_fee_status == 'partial':
        students = [s for s in students
                    if s.get('paid_fee', 0) > 0 and s.get('balance_fee', 0) > 0]

    all_classes = sorted([c for c in students_col.distinct('class') if c])
    all_sections = sorted([s for s in students_col.distinct('section') if s])

    return render_template('manage_students.html',
                         students=students,
                         all_classes=all_classes,
                         all_sections=all_sections,
                         filter_class=filter_class,
                         filter_section=filter_section,
                         filter_search=filter_search,
                         filter_fee_status=filter_fee_status)


@app.route('/admin/student/add', methods=['GET', 'POST'])
@admin_required
def add_student():
    if request.method == 'POST':
        name = SecurityValidator.sanitize_string(request.form.get('name', ''), 100)
        admission_no = SecurityValidator.sanitize_string(request.form.get('admission_no', ''), 20)
        roll_no = SecurityValidator.sanitize_string(request.form.get('roll_no', ''), 20)
        student_class = SecurityValidator.sanitize_string(request.form.get('class', ''), 20)
        section = SecurityValidator.sanitize_string(request.form.get('section', ''), 10)
        father_name = SecurityValidator.sanitize_string(request.form.get('father_name', ''), 100)
        mother_name = SecurityValidator.sanitize_string(request.form.get('mother_name', ''), 100)
        mobile = SecurityValidator.sanitize_string(request.form.get('mobile', ''), 15)
        address = SecurityValidator.sanitize_string(request.form.get('address', ''), 500)
        status = request.form.get('status', 'Active')

        if not name or not roll_no or not student_class:
            flash('⚠️ नाम, रोल नंबर और कक्षा अनिवार्य हैं!')
            return redirect(url_for('add_student'))

        valid, total_fee = SecurityValidator.validate_amount(request.form.get('total_fee', 0))
        if not valid:
            flash(f'⚠️ {total_fee}')
            return redirect(url_for('add_student'))

        if status not in ('Active', 'Inactive'):
            status = 'Active'

        student = {
            'name': name, 'admission_no': admission_no, 'roll_no': roll_no,
            'class': student_class, 'section': section,
            'father_name': father_name, 'mother_name': mother_name,
            'mobile': mobile, 'address': address,
            'total_fee': total_fee, 'paid_fee': 0, 'balance_fee': total_fee,
            'status': status,
            'added_at': get_ist_now_admin(),
            'added_by': session.get('admin_name', 'Admin')
        }
        students_col.insert_one(student)
        flash(f'✅ Student {name} सफलतापूर्वक जोड़ा गया!')
        return redirect(url_for('manage_students'))

    return render_template('add_student.html')


@app.route('/admin/student/edit/<student_id>', methods=['GET', 'POST'])
@admin_required
def edit_student(student_id):
    valid, _ = SecurityValidator.validate_object_id(student_id)
    if not valid:
        flash('Invalid student ID!')
        return redirect(url_for('manage_students'))

    student = students_col.find_one({'_id': ObjectId(student_id)})
    if not student:
        flash('Student नहीं मिला!')
        return redirect(url_for('manage_students'))

    if request.method == 'POST':
        valid, total_fee = SecurityValidator.validate_amount(request.form.get('total_fee', 0))
        if not valid:
            flash(f'⚠️ {total_fee}')
            return redirect(url_for('edit_student', student_id=student_id))

        paid_fee = student.get('paid_fee', 0)
        status = request.form.get('status', 'Active')
        if status not in ('Active', 'Inactive'):
            status = 'Active'

        updates = {
            'name': SecurityValidator.sanitize_string(request.form.get('name', ''), 100),
            'admission_no': SecurityValidator.sanitize_string(request.form.get('admission_no', ''), 20),
            'roll_no': SecurityValidator.sanitize_string(request.form.get('roll_no', ''), 20),
            'class': SecurityValidator.sanitize_string(request.form.get('class', ''), 20),
            'section': SecurityValidator.sanitize_string(request.form.get('section', ''), 10),
            'father_name': SecurityValidator.sanitize_string(request.form.get('father_name', ''), 100),
            'mother_name': SecurityValidator.sanitize_string(request.form.get('mother_name', ''), 100),
            'mobile': SecurityValidator.sanitize_string(request.form.get('mobile', ''), 15),
            'address': SecurityValidator.sanitize_string(request.form.get('address', ''), 500),
            'total_fee': total_fee,
            'balance_fee': total_fee - paid_fee,
            'status': status,
        }
        students_col.update_one({'_id': ObjectId(student_id)}, {'$set': updates})
        flash(f'✅ {updates["name"]} की जानकारी अपडेट हो गई!')
        return redirect(url_for('manage_students'))

    return render_template('edit_student.html', student=student)


@app.route('/admin/student/delete/<student_id>', methods=['POST'])
@admin_required
def delete_student(student_id):
    """Delete student — POST only."""
    valid, _ = SecurityValidator.validate_object_id(student_id)
    if not valid:
        flash('Invalid student ID!')
        return redirect(url_for('manage_students'))

    student = students_col.find_one({'_id': ObjectId(student_id)})
    if student:
        students_col.delete_one({'_id': ObjectId(student_id)})
        fee_history_col.delete_many({'student_id': str(student_id)})
        flash(f'🗑️ {student["name"]} और उनकी फीस हिस्ट्री हटा दी गई!')
    else:
        flash('Student नहीं मिला!')
    return redirect(url_for('manage_students'))


@app.route('/admin/student/pay/<student_id>', methods=['GET', 'POST'])
@admin_required
def pay_fee(student_id):
    valid, _ = SecurityValidator.validate_object_id(student_id)
    if not valid:
        flash('Invalid student ID!')
        return redirect(url_for('manage_students'))

    student = students_col.find_one({'_id': ObjectId(student_id)})
    if not student:
        flash('Student नहीं मिला!')
        return redirect(url_for('manage_students'))

    if request.method == 'POST':
        fee_fields = [
            ('reg_fee', 'Registration Fee'), ('form_charge', 'Form Charge'),
            ('prev_dues', 'Previous Dues'), ('tuition_fee', 'Tuition Fee'),
            ('computer_fee', 'Computer Fee'), ('admission_fee', 'Admission'),
            ('term_fee', 'Term Fee'), ('library_fee', 'Library Fee'),
            ('electric_charge', 'Electric Charge'),
            ('development_charge', 'Development Charge'),
            ('security_money', 'Security Money'),
            ('transport_fee', 'Conveyance/Transportation Fee'),
            ('exam_fee', 'Exam. Fee'), ('hostel_charge', 'Hostel Charge'),
            ('late_fine', 'Late Fine'), ('others_fee', 'Others'),
        ]

        breakdown = {}
        for field_name, label in fee_fields:
            raw = request.form.get(field_name, 0) or 0
            try:
                val = float(raw)
                if val < 0:
                    flash(f'⚠️ {label} negative नहीं हो सकती!')
                    return redirect(url_for('pay_fee', student_id=student_id))
                breakdown[label] = val
            except (ValueError, TypeError):
                breakdown[label] = 0.0

        amount = sum(breakdown.values())
        month = SecurityValidator.sanitize_string(request.form.get('month', ''), 50)
        payment_mode = SecurityValidator.sanitize_string(request.form.get('payment_mode', 'Cash'), 20)
        remarks = SecurityValidator.sanitize_string(request.form.get('remarks', ''), 500)

        if amount <= 0:
            flash('⚠️ कुल राशि 0 से अधिक होनी चाहिए!')
            return redirect(url_for('pay_fee', student_id=student_id))

        new_paid = student.get('paid_fee', 0) + amount
        new_balance = student.get('total_fee', 0) - new_paid

        students_col.update_one(
            {'_id': ObjectId(student_id)},
            {'$set': {'paid_fee': new_paid, 'balance_fee': new_balance}}
        )

        ist_now = get_ist_now_admin()
        receipt_no = f"GVP-FEE-{ist_now.strftime('%Y%m%d%H%M%S')}-{str(student_id)[-4:]}"

        fee_history_col.insert_one({
            'student_id': str(student_id),
            'student_name': student['name'],
            'class': student.get('class', ''),
            'section': student.get('section', ''),
            'roll_no': student.get('roll_no', ''),
            'admission_no': student.get('admission_no', ''),
            'amount': amount,
            'breakdown': breakdown,
            'month': month,
            'payment_mode': payment_mode,
            'remarks': remarks,
            'receipt_no': receipt_no,
            'date': ist_now.strftime('%Y-%m-%d %H:%M:%S'),
            'collected_by': session.get('admin_name', 'Admin'),
            'total_fee': student.get('total_fee', 0),
            'total_paid_after': new_paid,
            'balance_after': new_balance,
        })

        flash(f'✅ ₹{amount:,.0f} की फीस {student["name"]} के लिए जमा हो गई!')
        return redirect(url_for('fee_receipt', receipt_no=receipt_no))

    return render_template('pay_fee.html', student=student)


@app.route('/admin/student/receipt/<receipt_no>')
@admin_required
def fee_receipt(receipt_no):
    receipt_no = SecurityValidator.sanitize_string(receipt_no, 50)
    receipt = fee_history_col.find_one({'receipt_no': receipt_no})
    if not receipt:
        flash('Receipt नहीं मिली!')
        return redirect(url_for('manage_students'))

    student = None
    valid, _ = SecurityValidator.validate_object_id(receipt.get('student_id', ''))
    if valid:
        student = students_col.find_one({'_id': ObjectId(receipt['student_id'])})
    return render_template('fee_receipt.html', receipt=receipt, student=student)


@app.route('/admin/student/fee-history/<student_id>')
@admin_required
def fee_history(student_id):
    valid, _ = SecurityValidator.validate_object_id(student_id)
    if not valid:
        flash('Invalid student ID!')
        return redirect(url_for('manage_students'))

    student = students_col.find_one({'_id': ObjectId(student_id)})
    if not student:
        flash('Student नहीं मिला!')
        return redirect(url_for('manage_students'))
    history = list(fee_history_col.find({'student_id': str(student_id)}).sort('date', -1))
    return render_template('student_fee_history.html', student=student, history=history)


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Student Portal & Certificates
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/student/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def student_login():
    if session.get('student_id'):
        return redirect(url_for('student_dashboard'))

    if request.method == 'POST':
        roll_no = safe_str(request.form.get('roll_no', ''), 20).strip()
        mobile = safe_str(request.form.get('password', ''), 20).strip()

        if not roll_no or not mobile:
            flash('कृपया Roll No और Mobile Number दर्ज करें!', 'error')
            return render_template('student_login.html')

        # Secure authentication — match roll_no AND mobile only
        student = students_col.find_one({
            'roll_no': roll_no,
            'mobile': mobile
        })

        if student:
            session.clear()
            session['student_id'] = str(student['_id'])
            session['student_name'] = student['name']
            session.permanent = True
            flash("Login successful", 'success')
            return redirect(url_for('student_dashboard'))
        else:
            flash("Invalid credentials. Use Roll No and Mobile Number.", 'error')

    return render_template('student_login.html')


@app.route('/student/dashboard')
@student_required
def student_dashboard():
    student = students_col.find_one({'_id': ObjectId(session['student_id'])})
    if not student:
        session.clear()
        return redirect(url_for('student_login'))

    certificate = certificates_col.find_one({'student_id': session['student_id']})
    return render_template('student_dashboard.html', student=student, certificate=certificate)


@app.route('/student/certificate/<cert_id>')
@student_required
def view_certificate(cert_id):
    valid, _ = SecurityValidator.validate_object_id(cert_id)
    if not valid:
        flash("Invalid certificate ID.", "error")
        return redirect(url_for('student_dashboard'))

    cert = certificates_col.find_one({'_id': ObjectId(cert_id)})
    if not cert:
        flash("Certificate not found.", "error")
        return redirect(url_for('student_dashboard'))

    # IDOR prevention: verify certificate belongs to logged-in student
    if cert.get('student_id') != session.get('student_id'):
        log_security_event('IDOR_ATTEMPT', session.get('student_id', 'unknown'),
                          f'Tried to view cert {cert_id}')
        flash("Access denied.", "error")
        return redirect(url_for('student_dashboard'))

    student = students_col.find_one({'_id': ObjectId(cert['student_id'])})
    return render_template('certificate.html', cert=cert, student=student)


@app.route('/admin/certificates')
@admin_required
def admin_certificates():
    students = list(students_col.find().sort("name", 1))
    all_certs = list(certificates_col.find())
    cert_map = {c['student_id']: c for c in all_certs}
    return render_template('admin_certificates.html', students=students, cert_map=cert_map)


@app.route('/admin/certificate/generate/<student_id>', methods=['POST'])
@admin_required
def generate_certificate(student_id):
    valid, _ = SecurityValidator.validate_object_id(student_id)
    if not valid:
        flash("Invalid student ID", "error")
        return redirect(url_for('admin_certificates'))

    student = students_col.find_one({'_id': ObjectId(student_id)})
    if not student:
        flash("Student not found", "error")
        return redirect(url_for('admin_certificates'))

    course_name = SecurityValidator.sanitize_string(
        request.form.get('course_name', 'Annual Curriculum'), 100
    )
    grade = SecurityValidator.sanitize_string(request.form.get('grade', 'A+'), 10)

    # Cryptographically secure certificate number
    cert_no = f"CERT-{secrets.token_hex(4).upper()}"

    cert_data = {
        'student_id': str(student['_id']),
        'student_name': student['name'],
        'course': course_name,
        'grade': grade,
        'issue_date': datetime.now(timezone.utc).strftime("%d-%m-%Y"),
        'certificate_no': cert_no,
        'issued_by': session.get('admin_name', 'Administrator')
    }

    certificates_col.update_one(
        {'student_id': str(student['_id'])},
        {'$set': cert_data},
        upsert=True
    )

    flash(f"Certificate generated for {student['name']}", "success")
    return redirect(url_for('admin_certificates'))


@app.route('/admin/certificate/delete/<student_id>', methods=['POST'])
@admin_required
def delete_certificate(student_id):
    """Delete certificate — POST only."""
    student_id = SecurityValidator.sanitize_string(student_id, 30)
    certificates_col.delete_one({'student_id': student_id})
    flash("Certificate revoked successfully.", "success")
    return redirect(url_for('admin_certificates'))


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES — Health Check
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/health')
@csrf.exempt
def health_check():
    """Health check endpoint for monitoring."""
    try:
        client.server_info()
        return jsonify({'status': 'healthy', 'database': 'connected'}), 200
    except Exception:
        return jsonify({'status': 'unhealthy', 'database': 'disconnected'}), 503


# ═══════════════════════════════════════════════════════════════════════════
# ERROR HANDLERS
# ═══════════════════════════════════════════════════════════════════════════

@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html',
                         error='403 - Access Forbidden',
                         message='आपके पास इस पेज को देखने की अनुमति नहीं है।'), 403

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html',
                         error='404 - Page Not Found',
                         message='यह पेज मौजूद नहीं है।'), 404

@app.errorhandler(429)
def ratelimit_exceeded(e):
    return render_template('error.html',
                         error='429 - Too Many Requests',
                         message='बहुत अधिक अनुरोध! कृपया कुछ समय बाद प्रयास करें।'), 429

@app.errorhandler(500)
def internal_error(e):
    app.logger.error(f'Internal error: {e}')
    return render_template('error.html',
                         error='500 - Internal Server Error',
                         message='कुछ गलत हो गया। कृपया बाद में पुन: प्रयास करें।'), 500


# ═══════════════════════════════════════════════════════════════════════════
# STARTUP
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    init_admin()
    init_accountant(db)  # Pass shared DB connection

    app.run(
        host='0.0.0.0',
        port=int(os.environ.get('PORT', 5000)),
        debug=app.config.get('DEBUG', False)
    )
